import io
import json
import uuid
from unittest.mock import patch

from openpyxl import load_workbook

from django.contrib.admin.sites import site as admin_site
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core import mail
from django.test import RequestFactory, TestCase, override_settings

from checkin.admin import ParticipantAdmin
from checkin.admin_views import OPERATIONS_GROUP_NAME
from checkin.models import Event, Participant
from checkin.services.announcement_export import build_announcement_file
from checkin.services.application_confirmation_export import build_application_confirmation_file
from checkin.services.event_excel_export import (
    find_duplicate_labels,
    mask_name,
    mask_phone,
    participants_for_tab,
    remarks_for,
    split_display_name,
)
from checkin.services.label_assign import GROUP_SIZE, FixedEntry, FreshEntry, assign_genre
from checkin.services.score_sheet_export import build_score_sheet_file
from checkin.services.sheet_sync import push_order_for_event


def mulberry32(seed: int):
    """TypeScript 버전 테스트와 같은 계열의 시드 고정 PRNG (파이썬으로 이식, 자체 검증용)."""
    state = {"t": seed}

    def rng() -> float:
        state["t"] = (state["t"] + 0x6D2B79F5) & 0xFFFFFFFF
        t = state["t"]
        t = ((t ^ (t >> 15)) * (t | 1)) & 0xFFFFFFFF
        t = (t + (((t ^ (t >> 7)) * (t | 61)) & 0xFFFFFFFF)) & 0xFFFFFFFF
        return ((t ^ (t >> 14)) & 0xFFFFFFFF) / 4294967296

    return rng


class AssignGenreTests(TestCase):
    """
    lib/labelAssign.ts (Next.js 버전) / Code.gs (Apps Script 버전)와 동일한 시나리오로
    검증합니다 — 세 구현이 전부 같은 규칙을 지키는지 교차 확인하는 목적도 있습니다.
    """

    def test_25_names_with_3_duplicates_across_seeds(self):
        names = [f"사람{i}" for i in range(22)] + ["홍길동", "홍길동", "홍길동"]
        fresh = [FreshEntry(id=f"p{i}", name=n) for i, n in enumerate(names)]

        for seed in range(20):
            result = assign_genre([], fresh, mulberry32(seed))
            self.assertEqual(len(result), len(fresh))

            by_group: dict[str, list[str]] = {}
            for p in fresh:
                by_group.setdefault(result[p.id].group, []).append(p.name)

            for g, names_in_group in by_group.items():
                self.assertLessEqual(len(names_in_group), GROUP_SIZE)
                self.assertLessEqual(names_in_group.count("홍길동"), 1)

                nums = [result[p.id].number for p in fresh if result[p.id].group == g]
                self.assertEqual(len(set(nums)), len(nums))

    def test_fresh_duplicate_of_existing_fixed_entry_avoids_that_group(self):
        existing = [
            FixedEntry(id="e1", name="김철수", group="A", number=1),
            FixedEntry(id="e2", name="이영희", group="A", number=2),
        ]
        fresh = [
            FreshEntry(id="n1", name="김철수"),
            FreshEntry(id="n2", name="박민수"),
            FreshEntry(id="n3", name="최수정"),
        ]
        for seed in range(20):
            result = assign_genre(existing, fresh, mulberry32(seed))
            self.assertNotEqual(result["n1"].group, "A")

    def test_11_people_split_into_2_groups(self):
        fresh = [FreshEntry(id=f"p{i}", name=f"이름{i}") for i in range(11)]
        result = assign_genre([], fresh, mulberry32(1))
        groups_used = {slot.group for slot in result.values()}
        self.assertEqual(len(groups_used), 2)

    def test_100_person_stress_no_duplicate_names_in_any_group(self):
        pool = ["김민준", "이서연", "박도윤", "최지우", "정하은", "강서준", "조수아", "윤예준", "장하윤", "임시우"]
        pool_rng = mulberry32(999)
        fresh = [FreshEntry(id=f"p{i}", name=pool[int(pool_rng() * len(pool))]) for i in range(100)]

        result = assign_genre([], fresh, mulberry32(12345))
        by_group: dict[str, list[str]] = {}
        for p in fresh:
            by_group.setdefault(result[p.id].group, []).append(p.name)

        for names_in_group in by_group.values():
            counts: dict[str, int] = {}
            for n in names_in_group:
                counts[n] = counts.get(n, 0) + 1
            for c in counts.values():
                self.assertLessEqual(c, 1)

        self.assertEqual(len(result), 100)


class ExportMaskingHelperTests(TestCase):
    """마스킹/비고 헬퍼는 세 내보내기 전부가 공유하는 로직이라(#30), 순수 함수
    수준에서 따로 검증해둔다."""

    def test_mask_name_with_dancer_name(self):
        self.assertEqual(mask_name("김철수/비보이스파크"), "김*수/비보이스파크")

    def test_mask_name_without_separator(self):
        self.assertEqual(mask_name("김철수"), "김*수")

    def test_mask_name_dancer_same_as_real(self):
        self.assertEqual(mask_name("김철수/김철수"), "김*수/김*수")

    def test_mask_name_strips_spaces_around_separator(self):
        # "김철수 / 비보이스파크"처럼 슬래시 앞뒤에 공백이 있어도 본명만
        # 정확히 마스킹되어야 한다 (공백이 real_name에 섞여 들어가면 안 됨).
        self.assertEqual(mask_name("김철수 / 비보이스파크"), "김*수/비보이스파크")

    def test_split_display_name_strips_only_separator_spaces(self):
        # 이름 자체에 포함된 공백(예: "나나미 헤이지")은 보존하고, 구분자
        # 앞뒤 공백만 제거한다.
        self.assertEqual(split_display_name("나나미 헤이지 / bition hi"), ("나나미 헤이지", "bition hi"))

    def test_split_display_name_no_separator(self):
        self.assertEqual(split_display_name("김철수"), ("김철수", ""))

    def test_split_display_name_extra_slash_kept_in_dancer_part(self):
        # 첫 번째 슬래시 기준으로만 나눠서, 댄서명에 슬래시가 더 들어가도
        # 본명 추출이 깨지지 않는다.
        self.assertEqual(split_display_name("김철수/비보이/스파크"), ("김철수", "비보이/스파크"))

    def test_mask_phone_standard_11_digits(self):
        self.assertEqual(mask_phone("010-1234-5678"), "010-****-5678")

    def test_mask_phone_non_standard_left_as_is(self):
        self.assertEqual(mask_phone("02-123-4567"), "02-123-4567")

    def test_remarks_for_flags_unpaid_and_unverified(self):
        p = Participant(entry_type="참가", payment_status="PENDING", verification_status="PENDING")
        self.assertEqual(remarks_for(p), "미입금/학적 인증 필요")

    def test_remarks_for_empty_when_all_clear(self):
        p = Participant(entry_type="참가", payment_status="PAID", verification_status="APPROVED")
        self.assertEqual(remarks_for(p), "")

    def test_remarks_for_viewer_ignores_verification(self):
        # 관람은 학적검수 대상이 아니라, 미승인 상태여도 비고에 안 뜬다.
        p = Participant(entry_type="관람", payment_status="PENDING", verification_status="PENDING")
        self.assertEqual(remarks_for(p), "미입금")


class ParticipantNameNormalizationTests(TestCase):
    """저장 시점에 name의 "본명 / 댄서명" 구분자 앞뒤 공백을 정리해서, 목록/
    상세/CSV 백업처럼 name을 그대로 읽는 곳들이 따로 파싱하지 않아도 깨끗하게
    나오게 한다."""

    def setUp(self):
        self.event = Event.objects.create(volume=98, name="정규화 테스트 회차")

    def test_save_trims_spaces_around_separator(self):
        p = Participant.objects.create(
            event=self.event, entry_type="참가", name="김시현 / bition", phone="010-0000-0000"
        )
        self.assertEqual(p.name, "김시현/bition")

    def test_save_preserves_internal_spaces_in_each_part(self):
        p = Participant.objects.create(
            event=self.event, entry_type="참가", name="나나미 헤이지 / bition hi", phone="010-0000-0000"
        )
        self.assertEqual(p.name, "나나미 헤이지/bition hi")

    def test_save_leaves_already_clean_name_untouched(self):
        p = Participant.objects.create(
            event=self.event, entry_type="참가", name="김철수/비보이스파크", phone="010-0000-0000"
        )
        self.assertEqual(p.name, "김철수/비보이스파크")

    def test_save_no_separator_only_strips_outer_whitespace(self):
        p = Participant.objects.create(
            event=self.event, entry_type="관람", name="  최유진  ", phone="010-0000-0000"
        )
        self.assertEqual(p.name, "최유진")


class EventExcelExportTests(TestCase):
    """엑셀 내보내기 3종을 공용 모듈(event_excel_export.py)로 합친 리팩터링(#30)이
    기존 동작(탭 구성/마스킹/라벨 유무 필터링)을 그대로 유지하는지 확인."""

    def setUp(self):
        self.event = Event.objects.create(volume=99, name="테스트 회차")
        # 라벨이 배정된 확정 참가자 — 공지용/점수표 탭에 나와야 한다.
        self.labeled = Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", genre="Breaking",
            name="김철수/비보이스파크", phone="010-1234-5678", school="국민대학교",
            academic_status="재학", label_group="A", label_number=1, label_code="A-1",
            payment_status="PAID", verification_status="APPROVED", checkin_status="CHECKED_IN",
        )
        # 라벨이 아직 없는 신청자 — 신청 확인용 명단에는 나오되, 공지용/점수표에는 빠져야 한다.
        self.unlabeled = Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", genre="Breaking",
            name="박준혁", phone="010-2222-3333", payment_status="PENDING",
        )
        self.viewer = Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="관람",
            name="최유진", phone="010-4444-5555", payment_status="PAID",
        )

    def _sheet_rows(self, wb, sheet_title):
        ws = wb[sheet_title]
        # 1행 제목, 2행 헤더, 3행부터 데이터.
        return [[c.value for c in row] for row in ws.iter_rows(min_row=3)]

    def test_announcement_excel_masks_and_excludes_unlabeled(self):
        filename, content = build_announcement_file(self.event)
        self.assertTrue(filename.endswith(".xlsx"))
        wb = load_workbook(filename=io.BytesIO(content))

        self.assertIn("브레이킹", wb.sheetnames)
        self.assertIn("관람", wb.sheetnames)

        rows = self._sheet_rows(wb, "브레이킹")
        self.assertEqual(len(rows), 1, "라벨 없는 참가자는 공지용 명단 탭에서 빠져야 함")
        self.assertEqual(rows[0][1], "김*수/비보이스파크")  # 이름 마스킹
        self.assertEqual(rows[0][2], "010-****-5678")  # 연락처 마스킹
        self.assertEqual(rows[0][5], "A-1")  # 순서(라벨 코드)

        viewer_rows = self._sheet_rows(wb, "관람")
        self.assertEqual(len(viewer_rows), 1)
        self.assertEqual(viewer_rows[0][1], "최*진")

    def test_score_sheet_excel_no_masking_and_marks_paid_checked_in(self):
        filename, content = build_score_sheet_file(self.event)
        wb = load_workbook(filename=io.BytesIO(content))
        rows = self._sheet_rows(wb, "브레이킹")
        self.assertEqual(len(rows), 1, "라벨 없는 참가자는 점수표 탭에서 빠져야 함")
        self.assertEqual(rows[0][1], "김철수/비보이스파크")  # 마스킹 없음
        self.assertEqual(rows[0][2], "010-1234-5678")
        self.assertEqual(rows[0][6], "O")  # 입금 여부
        self.assertEqual(rows[0][7], "O")  # 도착 여부

    def test_application_confirmation_includes_unlabeled_participants(self):
        filename, content = build_application_confirmation_file(self.event)
        wb = load_workbook(filename=io.BytesIO(content))
        rows = self._sheet_rows(wb, "브레이킹")
        names = {row[1] for row in rows}
        self.assertEqual(len(rows), 2, "라벨 유무와 무관하게 신청자 전원이 나와야 함")
        self.assertIn("김*수/비보이스파크", names)
        self.assertIn("박*혁", names)  # 라벨 없는 신청자도 포함, 마스킹은 그대로 적용
        # "순서" 컬럼이 없는 게 이 내보내기의 특징 — 헤더 개수로 확인.
        self.assertEqual(len(rows[0]), 6)

    def test_find_duplicate_labels_empty_when_no_conflict(self):
        self.assertEqual(find_duplicate_labels(self.event), [])


class AdminSecurityRegressionTests(TestCase):
    """코드 리뷰(2026-08-30, PR #61)에서 발견된 실제 버그들에 대한 회귀 테스트.
    전부 실제로 재현/수정을 확인한 것들이라, 나중에 누가 관련 코드를 다시
    건드릴 때 조용히 되돌아가지 않도록 여기 고정해둔다."""

    def setUp(self):
        User = get_user_model()
        self.superuser = User.objects.create_superuser("root", "root@example.com", "pass12345")
        self.ops_user = User.objects.create_user("ops1", email="ops1@example.com", password="x", is_staff=True)
        Group.objects.get_or_create(name=OPERATIONS_GROUP_NAME)[0].user_set.add(self.ops_user)
        self.staff_user = User.objects.create_user("staff1", email="staff1@example.com", password="x", is_staff=True)
        self.target = User.objects.create_user("kim", email="kim@example.com", password="x", is_staff=True)

    def test_add_user_flow_does_not_crash(self):
        # AccountUserAdmin.save_related()가 add_form(role 필드 없음)에서도
        # cleaned_data["role"]을 무조건 읽어서 "+ 계정 초대"가 KeyError로
        # 500이 났었다.
        self.client.login(username="root", password="pass12345")
        resp = self.client.post("/admin/auth/user/add/", {
            "username": "newstaff", "password1": "Xk8f2m9qLp!", "password2": "Xk8f2m9qLp!",
        }, follow=True)
        self.assertEqual(resp.status_code, 200)
        User = get_user_model()
        self.assertTrue(User.objects.filter(username="newstaff").exists())

    def test_password_reset_requires_superuser(self):
        # send_password_reset이 admin_view()(=로그인한 스태프)만 확인해서,
        # 최하위 권한 스태프도 아무 계정에나 재설정 메일을 발송시킬 수 있었다.
        self.client.login(username="staff1", password="x")
        resp = self.client.post(f"/admin/auth/user/{self.target.pk}/send-password-reset/")
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(len(mail.outbox), 0)

    def test_password_reset_works_for_superuser(self):
        self.client.login(username="root", password="pass12345")
        resp = self.client.post(f"/admin/auth/user/{self.target.pk}/send-password-reset/", follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(mail.outbox), 1)

    def test_ops_user_cannot_self_promote_to_superuser(self):
        # auth.change_user 권한이 "운영진" 그룹에 부여되는 미래 상황을 가정 —
        # 그래도 role=super는 저장 시점에 서버에서 막혀야 한다.
        from django.contrib.auth.models import Permission
        from django.contrib.contenttypes.models import ContentType

        User = get_user_model()
        ct = ContentType.objects.get_for_model(User)
        perm = Permission.objects.get(content_type=ct, codename="change_user")
        self.ops_user.user_permissions.add(perm)

        self.client.login(username="ops1", password="x")
        resp = self.client.post(f"/admin/auth/user/{self.ops_user.pk}/change/", {
            "first_name": "", "email": "ops1@example.com", "role": "super", "is_active": "on",
        }, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.ops_user.refresh_from_db()
        self.assertFalse(self.ops_user.is_superuser)
        self.assertTrue(self.ops_user.groups.filter(name=OPERATIONS_GROUP_NAME).exists())

    def test_viewer_registration_leaves_genre_blank(self):
        # RegisterForm.genre에 빈 선택지가 없어서, 관람 신청자가 장르를 안
        # 건드려도 브라우저가 첫 옵션(Waacking)을 자동 제출해버렸다.
        Event.objects.create(volume=1, name="테스트", is_active=True)
        resp = self.client.post("/register/", {
            "entry_type": "관람", "name": "박관람", "phone": "010-0000-0000",
            "school": "국민대", "academic_status": "재학",
        })
        self.assertEqual(resp.status_code, 200)
        p = Participant.objects.get(name="박관람")
        self.assertIn(p.genre, (None, ""))

    def test_export_action_with_deleted_event_does_not_crash(self):
        # _first_selected_event가 queryset이 비어도(회차가 그 사이 삭제된
        # 경우) None 체크 없이 event.name 등을 바로 써서 500이 났었다.
        self.client.login(username="root", password="pass12345")
        event = Event.objects.create(volume=2, name="삭제될 회차")
        pk = event.pk
        event.delete()
        resp = self.client.post("/admin/checkin/event/", {
            "action": "export_announcement_excel", "_selected_action": [str(pk)],
        }, follow=True)
        self.assertEqual(resp.status_code, 200)

    def test_delete_selected_action_visible(self):
        # actions.html이 event/participant 액션 바를 통째로 하드코딩된
        # 버튼 몇 개로 바꾸면서, 장고 기본 "삭제" 액션이 화면에서 사라졌었다.
        self.client.login(username="root", password="pass12345")
        Event.objects.create(volume=3, name="목록에 보일 회차")
        html = self.client.get("/admin/checkin/event/").content.decode("utf-8")
        self.assertIn('value="delete_selected"', html)


class MarkRefundActionTests(TestCase):
    """환불 처리 액션 — 입금 완료(PAID) 상태만 환불 대상이 되고, 이미 발급된
    라벨/QR도 함께 회수되는지 (환불 후에도 옛 QR로 체크인 가능한 상태로
    남으면 안 됨)."""

    def setUp(self):
        User = get_user_model()
        self.superuser = User.objects.create_superuser("root", "root@example.com", "pass12345")
        self.event = Event.objects.create(volume=1, name="테스트 회차", is_active=True)
        self.client.login(username="root", password="pass12345")

    def _post_refund(self, participant_ids):
        return self.client.post("/admin/checkin/participant/", {
            "action": "mark_refund", "_selected_action": [str(pid) for pid in participant_ids],
        }, follow=True)

    def test_refund_revokes_existing_label_and_qr(self):
        p = Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", name="김환불", phone="010-0000-0001",
            genre="Breaking", verification_status="APPROVED", payment_status="PAID",
            label_group="A", label_number=1, label_code="A-1", qr_token=uuid.uuid4(),
        )
        self._post_refund([p.pk])
        p.refresh_from_db()
        self.assertEqual(p.payment_status, "REFUND")
        self.assertIsNone(p.label_group)
        self.assertIsNone(p.label_number)
        self.assertIsNone(p.label_code)
        self.assertIsNone(p.qr_token)

    def test_refund_ignores_non_paid_participants(self):
        # 대기 상태인 사람까지 실수로 같이 선택해도 잘못 바뀌면 안 된다.
        p = Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", name="이대기", phone="010-0000-0002",
            genre="Breaking", verification_status="PENDING", payment_status="PENDING",
        )
        self._post_refund([p.pk])
        p.refresh_from_db()
        self.assertEqual(p.payment_status, "PENDING")

    def test_refunded_participant_not_reissued_label_or_qr(self):
        # assign_labels_and_tokens()를 다시 돌려도 환불된 사람은 여전히
        # payment_status != PAID이므로 새 라벨/QR을 받지 않아야 한다.
        from checkin.services.assign_labels import assign_labels_and_tokens

        p = Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", name="박환불", phone="010-0000-0003",
            genre="Breaking", verification_status="APPROVED", payment_status="PAID",
            label_group="A", label_number=1, label_code="A-1", qr_token=uuid.uuid4(),
        )
        self._post_refund([p.pk])
        assign_labels_and_tokens(self.event)
        p.refresh_from_db()
        self.assertIsNone(p.label_code)
        self.assertIsNone(p.qr_token)


class ParticipantStatTilesTests(TestCase):
    """참가자 목록 상단 "학적검수 대기 및 반려" 타일 — 대기 상태뿐 아니라
    반려된 사람도 함께 세어야 한다(운영진이 둘 다 후속 조치가 필요한
    사람들이라 하나로 합쳐 보고 싶어함)."""

    def setUp(self):
        User = get_user_model()
        self.superuser = User.objects.create_superuser("root", "root@example.com", "pass12345")
        self.event = Event.objects.create(volume=1, name="테스트 회차", is_active=True)
        self.client.login(username="root", password="pass12345")

    def test_pending_and_rejected_both_counted(self):
        Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", name="대기자", phone="010-0000-0001",
            verification_status="PENDING", payment_status="PENDING",
        )
        Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", name="반려자", phone="010-0000-0002",
            verification_status="REJECTED", payment_status="PENDING",
        )
        Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", name="승인자", phone="010-0000-0003",
            verification_status="APPROVED", payment_status="PAID",
        )
        resp = self.client.get("/admin/checkin/participant/")
        self.assertEqual(resp.context["dbbt_stat_pending_verification"], 2)


class CheckinConfirmIdempotencyTests(TestCase):
    """체크인 확정 API를 두 번 불러도 최초 체크인 시각이 덮어써지지 않는지
    (#19) — _mark_checked_in을 select_for_update로 감싼 뒤(#85)에도 이
    동작이 그대로 유지되는지 확인."""

    def setUp(self):
        User = get_user_model()
        self.staff = User.objects.create_user("staff1", email="staff1@example.com", password="x", is_staff=True)
        self.event = Event.objects.create(volume=1, name="테스트 회차", is_active=True)
        self.participant = Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", name="김철수", phone="010-0000-0000",
            payment_status="PAID", verification_status="APPROVED",
        )
        self.client.login(username="staff1", password="x")

    def test_double_manual_checkin_keeps_first_checked_in_at(self):
        resp1 = self.client.post(f"/api/participants/{self.participant.pk}/manual-checkin/")
        self.assertEqual(resp1.status_code, 200)
        self.participant.refresh_from_db()
        first_time = self.participant.checked_in_at
        self.assertIsNotNone(first_time)
        self.assertEqual(self.participant.checkin_status, "CHECKED_IN")

        resp2 = self.client.post(f"/api/participants/{self.participant.pk}/manual-checkin/")
        self.assertEqual(resp2.status_code, 200)
        self.participant.refresh_from_db()
        self.assertEqual(self.participant.checked_in_at, first_time)


@override_settings(IMPORT_SECRET="test-secret")
class GoogleFormImportGenreValidationTests(TestCase):
    """구글 폼 연동(google_form_import)이 예비 신청 폼(RegisterForm)과 같은
    기준으로 참가자 장르를 요구하는지(#87) — 장르가 없거나 Genre에 없는
    값이면 저장하지 않고 errors로 알린다."""

    def setUp(self):
        self.event = Event.objects.create(volume=1, name="테스트 회차", is_active=True)

    def _import(self, rows):
        resp = self.client.post(
            "/api/import/google-form/",
            data=json.dumps({"rows": rows}),
            content_type="application/json",
            HTTP_X_IMPORT_SECRET="test-secret",
        )
        return resp.json()

    def test_participant_without_genre_is_rejected(self):
        data = self._import(
            [{"externalRef": "row1", "name": "김철수", "phone": "010-0000-0000", "type": "참가", "genre": ""}]
        )
        self.assertEqual(data["imported"], 0)
        self.assertEqual(len(data["errors"]), 1)
        self.assertFalse(Participant.objects.filter(external_ref="row1").exists())

    def test_participant_with_unknown_genre_is_rejected(self):
        # Genre.choices에 없는 값(예: 한글 표기, 오타) — GENRE_TABS 어떤 탭과도
        # 안 맞아서 그대로 저장하면 모든 엑셀에서 누락되는 "유령 참가자"가 된다.
        data = self._import(
            [{"externalRef": "row2", "name": "이영희", "phone": "010-0000-0001", "type": "참가", "genre": "왁킹"}]
        )
        self.assertEqual(data["imported"], 0)
        self.assertEqual(len(data["errors"]), 1)

    def test_participant_with_valid_genre_is_imported(self):
        data = self._import(
            [{"externalRef": "row3", "name": "박민수", "phone": "010-0000-0002", "type": "참가", "genre": "Waacking"}]
        )
        self.assertEqual(data["imported"], 1)
        self.assertEqual(Participant.objects.get(external_ref="row3").genre, "Waacking")

    def test_viewer_without_genre_is_still_imported(self):
        # 관람은 장르 문항 자체가 없으므로 검증 대상이 아니다.
        data = self._import([{"externalRef": "row4", "name": "최유진", "phone": "010-0000-0003", "type": "관람"}])
        self.assertEqual(data["imported"], 1)

    def test_viewer_with_genre_answered_as_viewer_choice_is_still_imported(self):
        # 실제 구글 폼은 "참가 장르" 문항에도 선택지로 "관람"이 있어서, 관람
        # 신청자는 그 문항에 "관람"이라고 답한 채로 들어온다(Genre.values에
        # 없는 값). "참가 / 관람" 문항이 필수라 entry_type은 항상 정확히
        # 판별되므로, genre는 참가자일 때만 검증되고 관람은 무조건 None으로
        # 지워져 이 값과 무관하게 정상 저장돼야 한다.
        data = self._import(
            [{"externalRef": "row5", "name": "정하윤", "phone": "010-0000-0004", "type": "관람", "genre": "관람"}]
        )
        self.assertEqual(data["imported"], 1)
        self.assertIsNone(Participant.objects.get(external_ref="row5").genre)


class SheetSyncOrderErrorHandlingTests(TestCase):
    """push_order_for_event가 OSError뿐 아니라 JSON이 아닌 응답(JSONDecodeError)도
    잡아서 관리자 화면에 500 대신 실패 메시지를 돌려주는지(#88) 확인."""

    def setUp(self):
        self.event = Event.objects.create(
            volume=1, name="테스트 회차", sheet_sync_url="https://example.com/exec"
        )
        Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", genre="Waacking",
            name="김철수", phone="010-0000-0000", label_group="A", label_number=1, label_code="A-1",
        )

    def test_non_json_response_returns_graceful_failure_instead_of_raising(self):
        # 웹 앱이 "나만" 액세스로 배포됐을 때 구글이 돌려주는 인증 안내
        # HTML(200 OK)을 흉내낸 응답.
        with patch("checkin.services.sheet_sync._post", return_value=b"<html>Authorization required</html>"):
            result = push_order_for_event(self.event)  # 예외 없이 끝나야 함
        self.assertFalse(result["ok"])
        self.assertIn("message", result)


class LabelGroupSortOrderTests(TestCase):
    """장르 하나가 GROUP_SIZE*26명을 넘어 두 글자 그룹("AA" 등, label_assign.py의
    _letter_at() 참고)이 생겨도, 명단/관리자 화면 정렬이 문자열 비교로 깨지지
    않고 A..Z 다음에 AA..가 오는지(#89) 확인."""

    def setUp(self):
        self.event = Event.objects.create(volume=1, name="테스트 회차")

    def _make(self, ref, group, number):
        return Participant.objects.create(
            id=uuid.uuid4(), event=self.event, entry_type="참가", genre="Waacking",
            name=f"참가자{ref}", phone=f"010-0000-{ref:04d}",
            label_group=group, label_number=number, label_code=f"{group}-{number}",
        )

    def test_participants_for_tab_orders_single_letter_before_double_letter(self):
        self._make(1, "AA", 1)
        self._make(2, "B", 1)
        self._make(3, "Z", 1)
        ordered = participants_for_tab(self.event, "Waacking")
        self.assertEqual([p.label_group for p in ordered], ["B", "Z", "AA"])

    def test_admin_label_sort_orders_single_letter_before_double_letter(self):
        self._make(1, "AA", 1)
        self._make(2, "B", 1)
        self._make(3, "Z", 1)
        request = RequestFactory().get("/admin/checkin/participant/")
        qs = ParticipantAdmin(Participant, admin_site).get_queryset(request).order_by("_label_sort")
        self.assertEqual([p.label_group for p in qs], ["B", "Z", "AA"])


@override_settings(AXES_ENABLED=True)
class LoginBruteForceProtectionTests(TestCase):
    """로그인 무차별 대입 방어(SEC-02, django-axes) — AXES_FAILURE_LIMIT(5회)만큼
    틀리면 그다음부터는 올바른 비밀번호를 넣어도 잠겨서 로그인이 안 되는지 확인.
    settings.py에서 manage.py test 실행 중에는 AXES_ENABLED를 꺼두므로(테스트
    클라이언트의 client.login()이 axes가 요구하는 request 없이 authenticate()를
    호출해 에러가 나기 때문), 이 테스트만 override_settings로 다시 켠 뒤
    client.login() 대신 실제 로그인 폼(POST /admin/login/)을 그대로 흉내내
    request를 정상적으로 넘긴다."""

    def setUp(self):
        User = get_user_model()
        self.superuser = User.objects.create_superuser("root", "root@example.com", "CorrectHorse123!")

    def _attempt(self, password):
        # 실제 로그인 폼도 next 히든 필드를 같이 보낸다 — 없으면 로그인
        # 성공 시 이 프로젝트에 없는 기본 리다이렉트(/accounts/profile/)로
        # 빠져 404가 나서, 테스트가 axes와 무관한 이유로 깨진다.
        return self.client.post(
            "/admin/login/", {"username": "root", "password": password, "next": "/admin/"}, follow=True
        )

    def test_locked_out_after_failure_limit_even_with_correct_password(self):
        # 처음 4번은 그냥 "아이디/비밀번호 틀림" 화면(200)만 다시 보여준다.
        for _ in range(4):
            resp = self._attempt("wrong-password")
            self.assertEqual(resp.status_code, 200)
            self.assertFalse(resp.wsgi_request.user.is_authenticated)

        # 5번째 실패로 AXES_FAILURE_LIMIT(5)에 도달하는 순간, 그 요청 자체가
        # 이미 잠금 응답(429 — AXES_HTTP_RESPONSE_CODE 기본값)으로 처리된다.
        resp = self._attempt("wrong-password")
        self.assertEqual(resp.status_code, 429)

        # 잠긴 뒤에는 올바른 비밀번호를 넣어도 여전히 막힌다.
        resp = self._attempt("CorrectHorse123!")
        self.assertEqual(resp.status_code, 429)
        self.assertFalse(resp.wsgi_request.user.is_authenticated)

    def test_correct_password_still_works_before_limit_reached(self):
        for _ in range(4):
            self._attempt("wrong-password")

        resp = self._attempt("CorrectHorse123!")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.wsgi_request.user.is_authenticated)


class AccountDeactivationRevokesActiveSessionTests(TestCase):
    """공용 기기 세션 관련 방어(SEC-04). 두 가지를 확인한다:
    1) 로그인 세션 유효기간이 Django 기본값(2주)이 아니라 7일로 줄었는지.
    2) "계정 활성화"를 끄면 다음 로그인부터가 아니라, 이미 로그인돼 있던
       세션도 바로 다음 요청부터 차단되는지 — 기기 분실 시 이 즉시 차단이
       실제 대응 수단으로 문서화(README, is_active 도움말)돼 있으므로,
       그 전제가 되는 동작 자체를 회귀 테스트로 고정해둔다."""

    def setUp(self):
        User = get_user_model()
        self.staff = User.objects.create_user("staff1", email="staff1@example.com", password="x", is_staff=True)

    def test_session_cookie_age_shortened_from_default(self):
        from django.conf import settings

        self.assertEqual(settings.SESSION_COOKIE_AGE, 60 * 60 * 24 * 7)
        self.assertLess(settings.SESSION_COOKIE_AGE, 1209600)  # Django 기본값(2주)보다 짧아야 함

    def test_deactivating_account_blocks_already_logged_in_session_immediately(self):
        self.client.force_login(self.staff)

        # 로그인된 상태에서는 스캐너 화면 접근 가능.
        resp = self.client.get("/checkin/")
        self.assertEqual(resp.status_code, 200)

        # 같은 세션(로그아웃 없이)인 채로, 다른 곳(관리자)이 계정을 비활성화.
        self.staff.is_active = False
        self.staff.save(update_fields=["is_active"])

        # 이 기기는 로그아웃한 적이 없는데도, 다음 요청부터 바로 로그인 화면으로 밀려나야 한다.
        resp = self.client.get("/checkin/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/admin/login/", resp.url)


class AdminAppListOrderingTests(TestCase):
    """admin 홈 화면 앱 목록에서 django-axes(로그인 실패 기록 — 운영진이 평소에
    볼 일 없는 부가 기능)가 알파벳순으로 맨 위에 오지 않고 맨 뒤로 가는지 확인."""

    def test_axes_app_sorted_last(self):
        User = get_user_model()
        superuser = User.objects.create_superuser("root", "root@example.com", "pass12345")
        self.client.force_login(superuser)
        resp = self.client.get("/admin/")
        app_labels = [app["app_label"] for app in resp.context["app_list"]]
        self.assertGreater(len(app_labels), 1, "다른 앱이 있어야 순서 비교가 의미 있음")
        self.assertEqual(app_labels[-1], "axes")
