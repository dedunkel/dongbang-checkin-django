"""
신청 참가자 확인용 공지 엑셀 다운로드.

기존 "공지용 명단"(announcement_export.py)은 라벨(예선 순서)이 배정된
확정 참가자만 담는 반면, 이건 라벨 배정 여부와 무관하게 그 장르에 신청한
사람 전원을 담아서, 참가자 본인이 자기 신청 내용이 맞게 접수됐는지
확인하는 용도다. 그래서 "순서" 컬럼이 없고, 아직 신청 처리가 덜 끝난
사람은 비고에 "미입금"/"학적 인증 필요"가 표시된다(문제없으면 빈칸).

이름/연락처 마스킹, 탭 구성, 제목 행 형식은 공지용 명단과 동일 — 관련
로직을 그대로 재사용한다.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from checkin.models import Event, Participant
from checkin.services.announcement_export import (
    GENRE_TABS,
    mask_name,
    mask_phone,
    remarks_for,
    write_title_row,
)

HEADERS = ["번호", "이름", "연락처", "소속대학", "학적", "비고"]


def participants_for_tab(event: Event, genre: str | None) -> list[Participant]:
    """라벨 유무와 무관하게 그 탭(장르 또는 관람)에 신청한 사람 전원을
    신청 순서(생성일시)로 정렬해 반환."""
    if genre is None:
        return list(Participant.objects.filter(event=event, entry_type="관람").order_by("created_at"))
    return list(
        Participant.objects.filter(event=event, entry_type="참가", genre=genre).order_by("created_at")
    )


def _write_sheet(wb: Workbook, event: Event, sheet_title: str, participants: list[Participant]) -> None:
    ws = wb.create_sheet(title=sheet_title)
    write_title_row(ws, event, sheet_title, len(HEADERS))

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    for i, p in enumerate(participants, start=1):
        ws.append(
            [
                i,
                mask_name(p.name),
                mask_phone(p.phone),
                p.school or "",
                p.academic_status or "",
                remarks_for(p),
            ]
        )

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_application_confirmation_workbook(event: Event) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for genre, tab_name in GENRE_TABS:
        _write_sheet(wb, event, tab_name, participants_for_tab(event, genre))

    return wb


def build_application_confirmation_file(event: Event) -> tuple[str, bytes]:
    """(파일명, xlsx 바이트) 튜플을 반환."""
    wb = build_application_confirmation_workbook(event)
    buf = io.BytesIO()
    wb.save(buf)
    filename = f"DongbangBattle Vol.{event.volume} 참가·관람 신청 확인용 명단.xlsx"
    return filename, buf.getvalue()
