"""
인포/공연장에서 쓰는 점수표 엑셀 다운로드.

공지용 명단(announcement_export.py)과 달리 이건 스태프 내부용이라 이름/
연락처를 마스킹하지 않는다. 입금 여부·도착 여부는 현재 DB 상태를 그대로
반영해서 미리 채워 넣고, 점수는 당연히 비워둔다(행사 당일 채워 넣는 용도).
장르별 탭 구성은 공지용 명단과 동일.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from checkin.models import Event
from checkin.services.announcement_export import GENRE_TABS, participants_for_tab, write_title_row

HEADERS = ["번호", "이름", "연락처", "소속대학", "학적", "순서", "입금 여부", "도착 여부", "점수"]

CHECK_MARK = "O"


def _write_sheet(wb: Workbook, event: Event, sheet_title: str, participants: list) -> None:
    ws = wb.create_sheet(title=sheet_title)
    write_title_row(ws, event, sheet_title, len(HEADERS))

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    for i, p in enumerate(participants, start=1):
        ws.append(
            [
                i,
                p.name,
                p.phone,
                p.school or "",
                p.academic_status or "",
                p.label_code or "",
                CHECK_MARK if p.payment_status == "PAID" else "",
                CHECK_MARK if p.checkin_status == "CHECKED_IN" else "",
                "",
            ]
        )

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_score_sheet_workbook(event: Event) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for genre, tab_name in GENRE_TABS:
        _write_sheet(wb, event, tab_name, participants_for_tab(event, genre))

    return wb


def build_score_sheet_file(event: Event) -> tuple[str, bytes]:
    """(파일명, xlsx 바이트) 튜플을 반환."""
    wb = build_score_sheet_workbook(event)
    buf = io.BytesIO()
    wb.save(buf)
    filename = f"DongbangBattle Vol.{event.volume} 점수표.xlsx"
    return filename, buf.getvalue()
