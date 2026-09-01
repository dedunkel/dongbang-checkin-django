"""
회차 엑셀 내보내기 3종(공지용 명단/신청 확인용 명단/점수표)이 공유하는 로직.

세 내보내기 전부 "장르별 탭 + 관람 탭 하나" 구성, 탭 맨 위 제목 행, 헤더 행,
컬럼 너비 지정이 구조적으로 동일하고 헤더 목록·행 값·필터링 규칙만 다르다
(#30). 예전에는 이 공용 로직이 "공지용" 전용이어야 할 announcement_export.py에
얹혀있어서 다른 두 모듈이 소유권이 맞지 않는 모듈을 그대로 import해 썼는데,
공지용만을 위한 변경이 점수표/신청확인용 export를 조용히 깨뜨릴 수 있는
구조였다 — 이 모듈로 분리해서 세 모듈이 동등하게 이 공용 헬퍼를 가져다 쓰게
했다.
"""

from __future__ import annotations

import io
import re
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from checkin.models import Event, Genre, Participant

# 장르별 탭 순서 — 관람은 마지막, genre=None으로 구분. 탭 목록은 Genre(모델)에서
# 그대로 가져온다 — 여기서 따로 문자열을 하드코딩하면 모델의 실제 선택지와
# 어긋날 수 있어(#48), 단일 소스로 유지한다.
GENRE_TABS = [(g.value, g.label) for g in Genre] + [(None, "관람")]

RowBuilder = Callable[[int, Participant], list]
TabParticipants = Callable[[Event, "str | None"], list[Participant]]

_PHONE_DIGITS_RE = re.compile(r"\D")


def _mask_middle(s: str) -> str:
    s = s.strip()
    if len(s) == 0:
        return s
    if len(s) == 1:
        return "*"
    if len(s) == 2:
        return s[0] + "*"
    if len(s) == 3:
        return s[0] + "*" + s[-1]
    # 4글자 이상은 글자 수에 비례해서 별표를 늘리지 않고, 앞 2글자만 남기고
    # 그다음 두 글자를 고정으로 별표 2개로 가린 뒤 나머지는 그대로 둔다
    # (예: "zaliij" -> "za**ij").
    return s[:2] + "**" + s[4:]


def split_display_name(raw_name: str) -> tuple[str, str]:
    """'이름 / 댄서명' 형식에서 구분자 앞뒤 공백만 제거하고 (본명, 댄서명)으로
    나눈다. 이름 자체에 포함된 공백(예: "나나미 헤이지")은 원형 그대로 둔다.
    구분자가 없으면 댄서명은 빈 문자열."""
    real, sep, dancer = raw_name.partition("/")
    return real.strip(), (dancer.strip() if sep else "")


def mask_name(raw_name: str) -> str:
    """'본명/댄서네임' 형식이면 본명 부분만 가운데를 마스킹하고 댄서네임은
    그대로 둔다. 구분자가 없거나(=댄서네임을 따로 안 적어 이름과 같은 경우),
    댄서네임을 본명과 똑같이 적은 경우(예: "김민준/김민준")는 양쪽 다
    동일하게 마스킹한다."""
    real, dancer = split_display_name(raw_name)
    masked_real = _mask_middle(real)
    if not dancer:
        return masked_real
    if dancer == real:
        return f"{masked_real}/{masked_real}"
    return f"{masked_real}/{dancer}"


def mask_phone(raw_phone: str) -> str:
    """010-****-1234 형식으로 통일. 11자리 숫자가 아니면 원본을 그대로 반환."""
    digits = _PHONE_DIGITS_RE.sub("", raw_phone or "")
    if len(digits) != 11:
        return raw_phone
    return f"{digits[:3]}-****-{digits[7:]}"


def remarks_for(p: Participant) -> str:
    """미입금/학적 인증 미승인 상태를 "미입금/학적 인증 필요" 식으로 합쳐서
    반환. 둘 다 문제없으면 빈 문자열."""
    notes = []
    if p.payment_status != "PAID":
        notes.append("미입금")
    if p.entry_type == "참가" and p.verification_status != "APPROVED":
        notes.append("학적 인증 필요")
    return "/".join(notes)


def participants_for_tab(event: Event, genre: str | None) -> list[Participant]:
    """탭 하나(장르 또는 관람)에 들어갈 "확정" 참가자를, 그 탭에서 보여줄
    순서로 정렬해 반환. 장르 탭은 라벨(조+번호)이 배정된 사람만, 라벨 코드
    기준으로 정렬. 관람 탭은 신청 순서(생성일시) 기준.

    라벨 배정 여부와 무관하게 신청자 전원이 필요한 경우(신청 확인용 명단)는
    이 함수를 쓰지 않는다 — application_confirmation_export.py의
    all_participants_for_tab 참고."""
    if genre is None:
        return list(Participant.objects.filter(event=event, entry_type="관람").order_by("created_at"))
    return sorted(
        Participant.objects.filter(event=event, entry_type="참가", genre=genre, label_code__isnull=False),
        key=lambda p: (p.label_group, p.label_number),
    )


def find_duplicate_labels(event: Event) -> list[str]:
    """장르 탭 안에서 같은 라벨 코드(예: "A-1")가 두 명 이상에게 붙어있는 경우를
    찾아 사람이 읽을 수 있는 문제 설명 목록으로 반환. 정상이면 빈 리스트.

    보통은 label_slot_unique DB 제약이 이런 중복 저장 자체를 막아주지만,
    (1) 그 제약이 못 잡는 예외 상황(예: genre가 비어있는 경우, NULL은 서로
    다르다고 취급돼서 제약에 안 걸림)이 있고, (2) 막히더라도 관리자에게는
    장르/코드/이름을 콕 집어 알려주는 게 훨씬 낫기 때문에 내보내기 직전에
    한 번 더 확인한다."""
    problems = []
    for genre, tab_name in GENRE_TABS:
        if genre is None:
            continue
        by_code: dict[str, list[str]] = {}
        for p in participants_for_tab(event, genre):
            by_code.setdefault(p.label_code, []).append(p.name)
        for code, names in by_code.items():
            if len(names) > 1:
                problems.append(f'{tab_name} 탭의 라벨 코드 "{code}"가 {len(names)}명에게 중복 배정됨: {", ".join(names)}')
    return problems


def write_title_row(ws, event: Event, sheet_title: str, num_columns: int) -> None:
    """탭 맨 위에 "동방배틀 Vol.N {장르} 참가자 명단"(관람은 "관람자 명단") 제목
    행을 병합 셀로 넣는다. 세 내보내기 전부 같은 제목 형식을 쓴다."""
    title_text = (
        f"동방배틀 Vol.{event.volume} 관람자 명단"
        if sheet_title == "관람"
        else f"동방배틀 Vol.{event.volume} {sheet_title} 참가자 명단"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")


def _write_sheet(
    wb: Workbook, event: Event, sheet_title: str, participants: list[Participant],
    headers: list[str], row_builder: RowBuilder,
) -> None:
    ws = wb.create_sheet(title=sheet_title)
    write_title_row(ws, event, sheet_title, len(headers))

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    for i, p in enumerate(participants, start=1):
        ws.append(row_builder(i, p))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_workbook(
    event: Event, headers: list[str], row_builder: RowBuilder,
    tab_participants: TabParticipants = participants_for_tab,
) -> Workbook:
    """탭 구성(장르별 + 관람)은 고정, 헤더/행 값/탭에 들어갈 참가자 선택
    규칙만 호출하는 쪽에서 넘겨받는 범용 워크북 빌더."""
    wb = Workbook()
    wb.remove(wb.active)  # 기본으로 생기는 빈 시트 제거

    for genre, tab_name in GENRE_TABS:
        _write_sheet(wb, event, tab_name, tab_participants(event, genre), headers, row_builder)

    return wb


def build_file(
    event: Event, headers: list[str], row_builder: RowBuilder, filename: str,
    tab_participants: TabParticipants = participants_for_tab,
) -> tuple[str, bytes]:
    """(파일명, xlsx 바이트) 튜플을 반환."""
    wb = build_workbook(event, headers, row_builder, tab_participants)
    buf = io.BytesIO()
    wb.save(buf)
    return filename, buf.getvalue()
